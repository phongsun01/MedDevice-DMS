import fitz
import docx
import openpyxl
import os

def extract_pdf(path):
    try:
        doc = fitz.open(path)
        text = ""
        for page in doc:
            text += page.get_text()
        return text
    except Exception as e:
        return f"Error PDF {path}: {e}"

def extract_docx(path):
    try:
        doc = docx.Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"Error DOCX {path}: {e}"

def extract_xlsx(path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text += "\t".join([str(c) for c in row if c is not None]) + "\n"
        return text
    except Exception as e:
        return f"Error XLSX {path}: {e}"

paths = {
    "Examion_PDF": r"d:\Antigravity\MedDeviceDMS\storage\files\thiet-bi-chan-doan-hinh-anh\x-quang\x-quang-examion\Examion_X-DR_BT_WS_brochure.pdf",
    "Examion_XLSX": r"d:\Antigravity\MedDeviceDMS\storage\files\thiet-bi-chan-doan-hinh-anh\x-quang\x-quang-examion\ĐUKT_MAY_XQUANG_ĐỨC_TẤM_PRUDENT_BẢN_CHUẨN.xlsx",
    "Fuji_PDF": r"d:\Antigravity\MedDeviceDMS\storage\files\thiet-bi-chan-doan-hinh-anh\x-quang\x-quang-fdr-68s\FDR_SMART f_Brochure- Vietnamese.pdf",
    "Fuji_DOCX": r"d:\Antigravity\MedDeviceDMS\storage\files\thiet-bi-chan-doan-hinh-anh\x-quang\x-quang-fdr-68s\TSKT FDR 68S - bàn 4 hướng.docx"
}

with open("extracted_specs.txt", "w", encoding="utf-8") as f:
    for name, path in paths.items():
        f.write(f"\n--- {name} ---\n")
        if path.lower().endswith(".pdf"):
            f.write(extract_pdf(path)[:8000])
        elif path.lower().endswith(".docx"):
            f.write(extract_docx(path)[:8000])
        elif path.lower().endswith(".xlsx"):
            f.write(extract_xlsx(path)[:8000])
print("Done saving to extracted_specs.txt")
