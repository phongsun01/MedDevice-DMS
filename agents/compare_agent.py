"""
MedDevice DMS - Compare Agent
Device specification comparison using Gemini AI + XLSX export.
"""
import json
from pathlib import Path

import structlog

from config import settings
from db import client as db

log = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# 1. Compare devices
# ---------------------------------------------------------------------------

async def compare_devices(
    device_id_a: str,
    device_id_b: str,
    telegram_user_id: str | None = None,
) -> dict:
    """Compare two devices. Prefer existing comparison docs, else use Gemini."""

    # Check existing comparison doc
    existing = await db.query(
        """SELECT * FROM document
           WHERE doc_type = 'comparison'
             AND (device = $a OR device = $b)
           LIMIT 1""",
        {"a": device_id_a, "b": device_id_b},
    )

    if existing and existing[0]:
        doc = existing[0][0] if isinstance(existing[0], list) else existing[0]
        if doc.get("content_text"):
            log.info("compare.existing_doc_found")
            comparison = _parse_existing_comparison(doc["content_text"])
            if comparison:
                return comparison

    # Fetch technical docs for both
    text_a = await _get_device_text(device_id_a)
    text_b = await _get_device_text(device_id_b)

    # Gemini extraction
    comparison = await _compare_with_gemini(text_a, text_b, device_id_a, device_id_b)

    # Audit log
    await db.create_audit_log(
        "compare", "device",
        f"{device_id_a} vs {device_id_b}",
        telegram_user_id,
        {"device_a": device_id_a, "device_b": device_id_b},
    )

    return comparison


async def _get_device_text(device_id: str) -> str:
    """Get the best technical document text for a device (prefer EN)."""
    results = await db.query(
        """SELECT content_text, sub_type FROM document
           WHERE device = $id AND doc_type = 'technical' AND content_text != NONE
           ORDER BY (sub_type = 'EN') DESC
           LIMIT 1""",
        {"id": device_id},
    )
    if results and results[0]:
        row = results[0][0] if isinstance(results[0], list) else results[0]
        return row.get("content_text", "")
    return ""


async def _compare_with_gemini(text_a: str, text_b: str, id_a: str, id_b: str) -> dict:
    """Use Gemini to extract and compare specifications."""
    from google import genai as google_genai

    client = google_genai.Client(api_key=settings.GEMINI_API_KEY)

    prompt = f"""Extract all technical specifications from the following two medical device documents as JSON.
Return format: {{"device_a": "name_a", "device_b": "name_b", "specs": [{{"name": "spec name", "value_a": "value", "value_b": "value"}}]}}

--- Document A (ID: {id_a}) ---
{text_a[:8000]}

--- Document B (ID: {id_b}) ---
{text_b[:8000]}
"""
    try:
        response = await client.aio.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
        )
        raw = response.text.strip().strip("```json").strip("```")
        return json.loads(raw)
    except BaseException as exc:
        log.error("compare.gemini_failed", error=str(exc))
        return {
            "device_a": str(id_a),
            "device_b": str(id_b),
            "specs": [],
            "error": str(exc),
        }


def _parse_existing_comparison(text: str) -> dict | None:
    """Try to parse an existing comparison document's text as JSON."""
    try:
        return json.loads(text)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 2. Render markdown table
# ---------------------------------------------------------------------------

def render_comparison_table_markdown(comparison: dict) -> str:
    """Render a Telegram-friendly comparison table."""
    dev_a = comparison.get("device_a", "Device A")
    dev_b = comparison.get("device_b", "Device B")
    specs = comparison.get("specs", [])

    if not specs:
        return "⚠️ Không có thông số để so sánh."

    lines = [
        f"📊 <b>So sánh thiết bị</b>",
        f"<b>{dev_a}</b> vs <b>{dev_b}</b>\n",
        f"{'Thông số':<25} | {dev_a[:12]:<14} | {dev_b[:12]}",
        "-" * 55,
    ]

    for spec in specs:
        name = spec.get("name", "")[:25]
        val_a = spec.get("value_a", "—")[:14]
        val_b = spec.get("value_b", "—")[:14]
        diff = " ⚠️" if val_a != val_b else ""
        lines.append(f"{name:<25} | {val_a:<14} | {val_b}{diff}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 3. Export XLSX
# ---------------------------------------------------------------------------

async def export_comparison_xlsx(comparison: dict, output_path: str) -> str:
    """Export comparison to a styled Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    dev_a = comparison.get("device_a", "Device A")
    dev_b = comparison.get("device_b", "Device B")

    # Header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Thông số kỹ thuật", dev_a, dev_b]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for i, spec in enumerate(comparison.get("specs", []), 2):
        ws.cell(row=i, column=1, value=spec.get("name", ""))
        ws.cell(row=i, column=2, value=spec.get("value_a", ""))
        ws.cell(row=i, column=3, value=spec.get("value_b", ""))

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("compare.xlsx_exported", path=output_path)
    return output_path


# ---------------------------------------------------------------------------
# 4. Handler (orchestrates everything)
# ---------------------------------------------------------------------------

async def compare_handler(
    device_name_a: str,
    device_name_b: str,
    telegram_user_id: str | None = None,
) -> tuple[str, str]:
    """Full compare flow: fuzzy match → compare → render → export."""
    from agents.search_agent import search_devices

    # Fuzzy match names to IDs
    results_a = await search_devices(device_name_a)
    results_b = await search_devices(device_name_b)

    if not results_a:
        return (f"❌ Không tìm thấy thiết bị: {device_name_a}", "")
    if not results_b:
        return (f"❌ Không tìm thấy thiết bị: {device_name_b}", "")

    id_a = results_a[0]["id"]
    id_b = results_b[0]["id"]

    comparison = await compare_devices(id_a, id_b, telegram_user_id)
    md_text = render_comparison_table_markdown(comparison)

    xlsx_path = f"storage/exports/compare_{id_a.split(':')[-1]}_{id_b.split(':')[-1]}.xlsx"
    await export_comparison_xlsx(comparison, xlsx_path)

    return (md_text, xlsx_path)
